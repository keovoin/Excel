#!/usr/bin/env python3
"""
Build a native Excel (.xlsx) sales dashboard from contoso_sales_export.csv.

Produces "Contoso_Sales_Dashboard.xlsx" with:
  - Dashboard sheet: KPI cards, summary tables, and native Excel charts
  - Data sheet: the full transaction table (formatted, filterable)

Run:  python3 build_dashboard.py
"""
import csv
from collections import defaultdict, OrderedDict

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

CSV_FILE = "contoso_sales_export.csv"
OUT_FILE = "Contoso_Sales_Dashboard.xlsx"

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# ---------- palette ----------
NAVY = "1A2342"
BLUE = "2B6CFF"
TEAL = "00B8A9"
GOLD = "F2C037"
PURPLE = "9B5DE5"
PINK = "F15BB5"
LIGHT = "F3F4F8"
WHITE = "FFFFFF"
INK = "1B1F2A"
MUTED = "6B7280"

thin = Side(style="thin", color="E7E9F0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


# ---------------------------------------------------------------- load
def load_rows():
    rows = []
    with open(CSV_FILE, newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            r["Units"] = int(r["Units"])
            r["Revenue"] = int(r["Revenue"])
            r["Profit"] = int(r["Profit"])
            r["Year"] = int(r["Date"][:4])
            r["MonthIdx"] = int(r["Date"][5:7]) - 1
            rows.append(r)
    return rows


def group_sum(rows, key, value):
    m = defaultdict(int)
    for r in rows:
        m[r[key]] += r[value]
    return OrderedDict(sorted(m.items(), key=lambda kv: kv[1], reverse=True))


# ---------------------------------------------------------------- styles helpers
def style_title(cell, text, size=18, color=INK, bold=True):
    cell.value = text
    cell.font = Font(name="Calibri", size=size, bold=bold, color=color)


def header_row(ws, row, start_col, headers, fill=NAVY):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = Font(bold=True, color=WHITE, size=11)
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER


def money(cell):
    cell.number_format = '$#,##0'


def pct(cell):
    cell.number_format = '0.0%'


# ---------------------------------------------------------------- build
def build():
    rows = load_rows()
    wb = Workbook()

    # ============================================================ DATA SHEET
    ws_data = wb.active
    ws_data.title = "Data"
    headers = ["Date", "Region", "Category", "Product", "Segment",
               "Units", "Revenue", "Profit"]
    ws_data.append(headers)
    for r in rows:
        ws_data.append([r["Date"], r["Region"], r["Category"], r["Product"],
                        r["Segment"], r["Units"], r["Revenue"], r["Profit"]])

    n = len(rows)
    last = n + 1
    for row in range(2, last + 1):
        money(ws_data.cell(row=row, column=7))
        money(ws_data.cell(row=row, column=8))

    tbl = Table(displayName="SalesData", ref=f"A1:H{last}")
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws_data.add_table(tbl)
    widths = [12, 16, 16, 24, 13, 8, 12, 12]
    for i, w in enumerate(widths, start=1):
        ws_data.column_dimensions[get_column_letter(i)].width = w
    ws_data.freeze_panes = "A2"

    # ============================================================ AGGREGATES
    total_rev = sum(r["Revenue"] for r in rows)
    total_prof = sum(r["Profit"] for r in rows)
    orders = n
    margin = total_prof / total_rev if total_rev else 0
    aov = total_rev / orders if orders else 0
    total_units = sum(r["Units"] for r in rows)

    by_cat = group_sum(rows, "Category", "Revenue")
    by_reg = group_sum(rows, "Region", "Revenue")
    by_seg = group_sum(rows, "Segment", "Revenue")
    by_prod = group_sum(rows, "Product", "Revenue")
    top_prod = OrderedDict(list(by_prod.items())[:8])

    # monthly trend across years -> ordered labels
    monthly = OrderedDict()
    years = sorted({r["Year"] for r in rows})
    for y in years:
        for m in range(12):
            monthly[(y, m)] = {"rev": 0, "prof": 0}
    for r in rows:
        k = (r["Year"], r["MonthIdx"])
        monthly[k]["rev"] += r["Revenue"]
        monthly[k]["prof"] += r["Profit"]

    # ============================================================ DASHBOARD SHEET
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False
    for col in range(1, 30):
        ws.column_dimensions[get_column_letter(col)].width = 11

    # Banner
    ws.merge_cells("B2:N2")
    style_title(ws["B2"], "  CONTOSO SALES  —  EXECUTIVE DASHBOARD", size=20, color=WHITE)
    for col in range(2, 15):
        ws.cell(row=2, column=col).fill = PatternFill("solid", fgColor=NAVY)
    ws.row_dimensions[2].height = 34
    ws.merge_cells("B3:N3")
    ws["B3"] = "  Sample data  •  generated dashboard  •  all figures are illustrative"
    ws["B3"].font = Font(italic=True, size=10, color=MUTED)

    # ---- KPI cards (row 5-7) ----
    kpis = [
        ("TOTAL REVENUE", total_rev, BLUE, '$#,##0'),
        ("TOTAL PROFIT", total_prof, TEAL, '$#,##0'),
        ("ORDERS", orders, GOLD, '#,##0'),
        ("PROFIT MARGIN", margin, PURPLE, '0.0%'),
        ("AVG ORDER VALUE", aov, PINK, '$#,##0'),
    ]
    col = 2
    for label, value, color, fmt in kpis:
        c1 = get_column_letter(col)
        c2 = get_column_letter(col + 1)
        ws.merge_cells(f"{c1}5:{c2}5")
        ws.merge_cells(f"{c1}6:{c2}6")
        lab = ws[f"{c1}5"]
        lab.value = label
        lab.font = Font(bold=True, size=9, color=WHITE)
        lab.alignment = Alignment(horizontal="center", vertical="center")
        val = ws[f"{c1}6"]
        val.value = value
        val.number_format = fmt
        val.font = Font(bold=True, size=20, color=WHITE)
        val.alignment = Alignment(horizontal="center", vertical="center")
        for rr in (5, 6):
            for cc in (col, col + 1):
                ws.cell(row=rr, column=cc).fill = PatternFill("solid", fgColor=color)
        ws.row_dimensions[5].height = 18
        ws.row_dimensions[6].height = 40
        col += 3

    # ---- Helper to write a small summary table ----
    def write_table(anchor_row, anchor_col, title, mapping, value_header="Revenue"):
        ac = get_column_letter(anchor_col)
        ws.cell(row=anchor_row, column=anchor_col, value=title).font = Font(bold=True, size=12, color=INK)
        hr = anchor_row + 1
        header_row(ws, hr, anchor_col, [title.split(" by ")[-1] if " by " in title else "Item", value_header])
        r = hr + 1
        for k, v in mapping.items():
            kc = ws.cell(row=r, column=anchor_col, value=k)
            vc = ws.cell(row=r, column=anchor_col + 1, value=v)
            money(vc)
            kc.border = BORDER
            vc.border = BORDER
            r += 1
        return hr, r - 1  # header row, last data row

    # ---- Table: Revenue by Category (B9) + Pie chart ----
    cat_hdr, cat_last = write_table(9, 2, "Revenue by Category", by_cat)
    pie = PieChart()
    pie.title = "Revenue by Category"
    labels = Reference(ws, min_col=2, min_row=cat_hdr + 1, max_row=cat_last)
    data = Reference(ws, min_col=3, min_row=cat_hdr, max_row=cat_last)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.height = 6.5
    pie.width = 11
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    ws.add_chart(pie, "E9")

    # ---- Table: Revenue by Region (B18) + Bar chart ----
    reg_hdr, reg_last = write_table(18, 2, "Revenue by Region", by_reg)
    bar = BarChart()
    bar.type = "col"
    bar.title = "Revenue by Region"
    bar.legend = None
    rdata = Reference(ws, min_col=3, min_row=reg_hdr, max_row=reg_last)
    rcats = Reference(ws, min_col=2, min_row=reg_hdr + 1, max_row=reg_last)
    bar.add_data(rdata, titles_from_data=True)
    bar.set_categories(rcats)
    bar.height = 6.5
    bar.width = 11
    bar.y_axis.numFmt = '$#,##0'
    ws.add_chart(bar, "E18")

    # ---- Table: Sales by Segment (B28) + Bar ----
    seg_hdr, seg_last = write_table(28, 2, "Revenue by Segment", by_seg)
    sbar = BarChart()
    sbar.type = "bar"
    sbar.title = "Revenue by Segment"
    sbar.legend = None
    sdata = Reference(ws, min_col=3, min_row=seg_hdr, max_row=seg_last)
    scats = Reference(ws, min_col=2, min_row=seg_hdr + 1, max_row=seg_last)
    sbar.add_data(sdata, titles_from_data=True)
    sbar.set_categories(scats)
    sbar.height = 6.5
    sbar.width = 11
    sbar.x_axis.numFmt = '$#,##0'
    ws.add_chart(sbar, "E28")

    # ---- Top products table (far right) + horizontal bar ----
    tp_hdr, tp_last = write_table(9, 16, "Top Products by Revenue", top_prod)
    ws.column_dimensions[get_column_letter(16)].width = 24
    pbar = BarChart()
    pbar.type = "bar"
    pbar.title = "Top 8 Products"
    pbar.legend = None
    pdata = Reference(ws, min_col=17, min_row=tp_hdr, max_row=tp_last)
    pcats = Reference(ws, min_col=16, min_row=tp_hdr + 1, max_row=tp_last)
    pbar.add_data(pdata, titles_from_data=True)
    pbar.set_categories(pcats)
    pbar.height = 8
    pbar.width = 13
    pbar.x_axis.numFmt = '$#,##0'
    ws.add_chart(pbar, "S9")

    # ---- Monthly trend table (hidden-ish area) + line chart ----
    trend_title_row = 40
    ws.cell(row=trend_title_row, column=2, value="Monthly Revenue & Profit Trend").font = Font(bold=True, size=12, color=INK)
    th = trend_title_row + 1
    header_row(ws, th, 2, ["Month", "Revenue", "Profit"])
    r = th + 1
    for (y, m), v in monthly.items():
        ws.cell(row=r, column=2, value=f"{MONTHS[m]} {str(y)[2:]}")
        money(ws.cell(row=r, column=3, value=v["rev"]))
        money(ws.cell(row=r, column=4, value=v["prof"]))
        for cc in (2, 3, 4):
            ws.cell(row=r, column=cc).border = BORDER
        r += 1
    trend_last = r - 1

    line = LineChart()
    line.title = "Revenue & Profit by Month"
    line.style = 12
    ldata = Reference(ws, min_col=3, max_col=4, min_row=th, max_row=trend_last)
    lcats = Reference(ws, min_col=2, min_row=th + 1, max_row=trend_last)
    line.add_data(ldata, titles_from_data=True)
    line.set_categories(lcats)
    line.height = 8
    line.width = 26
    line.y_axis.numFmt = '$#,##0'
    ws.add_chart(line, "F40")

    # nice column widths for product col on right
    ws.column_dimensions["P"].width = 24

    wb.save(OUT_FILE)
    print(f"Saved {OUT_FILE}")
    print(f"  rows={n}  revenue=${total_rev:,}  profit=${total_prof:,}  "
          f"margin={margin*100:.1f}%  units={total_units:,}")


if __name__ == "__main__":
    build()
